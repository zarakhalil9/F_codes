import hmac
import requests
import hashlib
import base64
import time
import uuid
import json
import openpyxl

# Constants
api_key = "guu7gs1XhGsCkq0YMcFsbXBandW8jUAu"
secret_key = "m0CO2odX602aMeqJSbQsgjqYGLsNBU1NgGQ7xLkeYhM"
url = "https://prod.emea.api.fiservapps.com/sandbox/ipp/payments-gateway/v2/payment-url"

# Function to generate message signature
def generate_message_signature(secret_key, message):
    signature = hmac.new(
        key=secret_key.encode(),
        msg=message.encode(),
        digestmod=hashlib.sha256
    ).digest()
    return base64.b64encode(signature).decode()

# Function to update the Excel file with payment URLs
def update_excel_with_payment_urls(file_path):
    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Find the header row and create a mapping for the order IDs
    headers = [cell.value for cell in ws[1]]
    order_id_index = headers.index('order_id')
    payment_url_index = headers.index('payment_url')
    
    for row in ws.iter_rows(min_row=2, values_only=False):
        order_id = row[order_id_index].value
        if not order_id:
            continue

        # Prepare the request body
        body = {
            "transactionAmount": {
                "total": 20.50,
                "currency": "AED"
            },
            "transactionType": "SALE",
            "transactionNotificationURL": "https://webhook.site/53df3d07-0f4d-403a-bc87-d7d23e8db966",
            "expiration": 5102358400,
            "orderId": order_id,  # Use actual order ID from the Excel file
            "invoiceNumber": "INV1234",
            "purchaseOrderNumber": "PO1042",
            "billing": {
                "name": "test 123",
                "customerId": "1234567890",
                "contact": {
                    "email": ""
                }
            }
        }

        # Generate unique client request ID and timestamp
        client_request_id = str(uuid.uuid4())
        timestamp = int(time.time() * 1000)

        # Create message string
        message = f"{api_key}{client_request_id}{timestamp}{json.dumps(body)}"
        message_signature = generate_message_signature(secret_key, message)

        # Define headers
        headers = {
            "Content-Type": "application/json",
            "Client-Request-Id": client_request_id,
            "Api-Key": api_key,
            "Timestamp": str(timestamp),
            "Message-Signature": message_signature
        }

        # Send POST request
        try:
            response = requests.post(url, headers=headers, json=body)
            response_data = response.json()

            # Check if the response is successful
            if response.status_code == 200 and response_data.get("requestStatus") == "SUCCESS":
                payment_url = response_data.get("paymentUrl")
                row[payment_url_index].value = payment_url
                print(f"Payment URL updated for order {order_id}: {payment_url}")
            else:
                print(f"Failed to get payment URL for order {order_id}: {response_data}")
        except requests.RequestException as e:
            print(f"Error making request for order {order_id}: {e}")

    # Save the workbook with updated URLs
    wb.save(file_path)
    print(f"Excel file updated and saved as {file_path}")

# Update the Excel file with payment URLs
file_path = r'C:\Users\adil_\OneDrive\Desktop\Automation\customer_data.xlsx'  # Update this path to your actual file
update_excel_with_payment_urls(file_path)

