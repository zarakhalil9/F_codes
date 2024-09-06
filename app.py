from flask import Flask, request, render_template, redirect, url_for, flash
import openpyxl
import pandas as pd
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.utils import formataddr
from werkzeug.utils import secure_filename
import os
from datetime import datetime
from apscheduler.schedulers.background import BackgroundScheduler
import requests
import hmac
import hashlib
import base64
import json
import uuid
import time

app = Flask(__name__)
app.secret_key = os.urandom(24)  # Securely generated secret key

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Initialize scheduler
scheduler = BackgroundScheduler()
scheduler.start()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file part', 'error')
            return redirect(request.url)
        file = request.files['file']
        schedule_time = request.form.get('schedule')

        if file.filename == '':
            flash('No selected file', 'error')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)

            if schedule_time:
                # Schedule the email sending
                schedule_datetime = datetime.fromisoformat(schedule_time)
                scheduler.add_job(
                    func=send_bulk_emails, 
                    trigger='date', 
                    run_date=schedule_datetime, 
                    args=[file_path]
                )
                flash('Emails scheduled successfully!', 'success')
            else:
                send_bulk_emails(file_path)
                flash('Emails sent successfully!', 'success')

            return redirect(url_for('display_file', filename=filename))
        else:
            flash('Invalid file format. Please upload an Excel file.', 'error')
            return redirect(request.url)
    return render_template('index.html')

@app.route('/display/<filename>')
def display_file(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    df = pd.read_excel(file_path)
    
    # Identify missing fields and create a summary DataFrame
    missing_data = df.isnull()
    missing_rows = missing_data[missing_data.any(axis=1)]
    
    # Create a summary DataFrame
    summary_df = pd.DataFrame({
        'Row': missing_rows.index + 2,  # +2 because DataFrame index starts at 0 and header is row 1
        'Field': [', '.join(missing_rows.columns[missing_data.loc[row]]) for row in missing_rows.index],
    })
    
    # Convert DataFrames to HTML
    df_html = df.to_html(classes='table table-bordered', na_rep='NaN', escape=False)
    summary_html = summary_df.to_html(classes='table table-bordered', index=False)
    
    return render_template('display.html', tables=[df_html, summary_html], filename=filename)

@app.route('/send_emails/<filename>', methods=['POST'])
def send_emails(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    errors = send_bulk_emails(file_path)
    
    if errors:
        for error in errors:
            flash(error, 'error')
    else:
        flash('Emails sent successfully!', 'success')
    
    return redirect(url_for('upload_file'))

@app.route('/generate_payment_urls/<filename>', methods=['POST'])
def generate_payment_urls(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    try:
        update_excel_with_payment_urls(file_path)
        flash('Payment URLs generated and updated successfully!', 'success')
    except Exception as e:
        flash(f'Error generating payment URLs: {e}', 'error')
    
    return redirect(url_for('display_file', filename=filename))

def generate_message_signature(secret_key, message):
    signature = hmac.new(
        key=secret_key.encode(),
        msg=message.encode(),
        digestmod=hashlib.sha256
    ).digest()
    return base64.b64encode(signature).decode()

def update_excel_with_payment_urls(file_path):
    # Define constants
    api_key = "guu7gs1XhGsCkq0YMcFsbXBandW8jUAu"
    secret_key = "m0CO2odX602aMeqJSbQsgjqYGLsNBU1NgGQ7xLkeYhM"
    url = "https://prod.emea.api.fiservapps.com/sandbox/ipp/payments-gateway/v2/payment-url"

    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Find the header row and create a mapping for the order IDs
    headers = [cell.value for cell in ws[1]]
    order_id_index = headers.index('order_id')
    payment_url_index = headers.index('payment_url')
    
    for row in ws.iter_rows(min_row=2, values_only=False):
        order_id = row[order_id_index].value
        if not order_id:
            continue

        # Prepare the request body
        body = {
            "transactionAmount": {
                "total": 20.50,
                "currency": "AED"
            },
            "transactionType": "SALE",
            "transactionNotificationURL": "https://webhook.site/53df3d07-0f4d-403a-bc87-d7d23e8db966",
            "expiration": 5102358400,
            "orderId": order_id,  # Use actual order ID from the Excel file
            "invoiceNumber": "INV1234",
            "purchaseOrderNumber": "PO1042",
            "billing": {
                "name": "test 123",
                "customerId": "1234567890",
                "contact": {
                    "email": ""
                }
            }
        }

        # Generate unique client request ID and timestamp
        client_request_id = str(uuid.uuid4())
        timestamp = int(time.time() * 1000)

        # Create message string
        message = f"{api_key}{client_request_id}{timestamp}{json.dumps(body)}"
        message_signature = generate_message_signature(secret_key, message)

        # Define headers
        headers = {
            "Content-Type": "application/json",
            "Client-Request-Id": client_request_id,
            "Api-Key": api_key,
            "Timestamp": str(timestamp),
            "Message-Signature": message_signature
        }

        # Send POST request
        try:
            response = requests.post(url, headers=headers, json=body)
            response_data = response.json()

            # Check if the response is successful
            if response.status_code == 200 and response_data.get("requestStatus") == "SUCCESS":
                payment_url = response_data.get("paymentUrl")
                row[payment_url_index].value = payment_url
                print(f"Payment URL updated for order {order_id}: {payment_url}")
            else:
                print(f"Failed to get payment URL for order {order_id}: {response_data}")
        except requests.RequestException as e:
            print(f"Error making request for order {order_id}: {e}")

    # Save the workbook with updated URLs
    wb.save(file_path)
    print(f"Excel file updated and saved as {file_path}")

def send_bulk_emails(excel_file_path):
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    smtp_user = ""
    smtp_password = ""

    df = pd.read_excel(excel_file_path)
    errors = []

    for _, row in df.iterrows():
        email = row['email']
        if pd.isna(email):
            continue  # Skip rows with NaN emails
        try:
            msg = MIMEMultipart()
            msg['From'] = formataddr(('Your Name', smtp_user))
            msg['To'] = email
            msg['Subject'] = "Your Payment Link"

            body = f"""Dear {row['customer_name']},

Here is your payment link for order {row['order_id']}.

Amount: {row['amount']} {row['currency']}
Transaction Type: {row['transaction_type']}
Expiry Date: {row['payment_expiry_date']}
Payment Link: {row['payment_url']}

Thank you!
"""
            msg.attach(MIMEText(body, 'plain'))

            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls()
                server.login(smtp_user, smtp_password)
                server.send_message(msg)
            print(f"Email sent to {email}")

        except Exception as e:
            errors.append(f"Failed to send email to {email}: {e}")

    return errors

if __name__ == "__main__":
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)
    app.run(debug=True)
